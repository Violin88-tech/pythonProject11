from openpyxl import load_workbook

workbook = load_workbook("tmp/~Pairwise.xlsx")
sheet = workbook.active
print(sheet.cell(raw=1, column=1).value)
